# app/core/generador_reportes.py
from __future__ import annotations
from typing import Dict, Tuple
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from fpdf import FPDF

# ---------------- Excel ----------------

def _fmt_currency(symbol: str) -> str:
    return f'"{symbol}"#,##0.00' if symbol else "#,##0.00"

def _write_df(ws: Worksheet, df: pd.DataFrame, header_fill=None, freeze=True, auto_filter=True):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for c in ws[1]:
                c.font = Font(bold=True, color="FFFFFFFF")
                if header_fill:
                    c.fill = header_fill
                c.alignment = Alignment(vertical="center")
    if freeze: ws.freeze_panes = "A2"
    if auto_filter and ws.max_row >= 2 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

def _apply_number_formats(ws: Worksheet, df: pd.DataFrame, currency_symbol: str):
    for j, col in enumerate(df.columns, start=1):
        col_l = str(col).lower()
        if "ingreso" in col_l or "monto" in col_l or "importe" in col_l:
            fmt = _fmt_currency(currency_symbol)
        elif "porcentaje" in col_l:
            fmt = "0.00%"
        elif "puntaje" in col_l:
            fmt = "0.00"
        elif "edad" in col_l or col_l == "id":
            fmt = "0"
        else:
            fmt = None
        if fmt:
            for cell in ws.iter_cols(min_col=j, max_col=j, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.number_format = fmt

def _sheet_simple(wb: Workbook, name: str, df: pd.DataFrame, color_hex: str, currency_symbol: str):
    ws = wb.create_sheet(name)
    _write_df(ws, df, header_fill=PatternFill("solid", fgColor=color_hex))
    _apply_number_formats(ws, df, currency_symbol)
    return ws

def _append_block(ws: Worksheet, obj):
    fill = PatternFill("solid", fgColor="FFEDE7D9")
    if isinstance(obj, pd.DataFrame) and not obj.empty:
        _write_df(ws, obj, header_fill=fill, freeze=False, auto_filter=True)
    else:
        ws.append(["OK"])

def generar_excel(
    df_original: pd.DataFrame,
    df_limpio: pd.DataFrame,
    df_resumen: pd.DataFrame,
    df_cat: pd.DataFrame,
    validaciones: dict,
    parametros: Dict[str, str],
    currency_symbol: str,
    salida: Path,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    _sheet_simple(wb, "datos_limpios", df_limpio, "FF2B3A55", currency_symbol)
    _sheet_simple(wb, "resumen", df_resumen, "FF2D4059", currency_symbol)
    _sheet_simple(wb, "por_categoria", df_cat, "FF3C4858", currency_symbol)

    ws4 = wb.create_sheet("Validaciones")
    ws4.append(["Duplicados por ID"]); _append_block(ws4, validaciones.get("duplicados_por_id"))
    ws4.append([]); ws4.append(["Edades fuera de rango (<0 o >120)"]); _append_block(ws4, validaciones.get("edades_fuera"))
    ws4.append([]); ws4.append(["Ingresos no válidos (<=0 o NaN)"]); _append_block(ws4, validaciones.get("ingresos_invalidos"))
    ws4.append([]); ws4.append(["Filas con NaN (muestra)"]); _append_block(ws4, validaciones.get("filas_nan"))

    ws5 = wb.create_sheet("Parámetros")
    ws5.append(["Parametro","Valor"])
    for k, v in parametros.items():
        ws5.append([k, v])

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)

# ---------------- PDF (Unicode con fallback) ----------------

def _try_add_dejavu(pdf: FPDF) -> Tuple[bool,str]:
    candidates = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/DejaVuSans.ttf",
        "app/static/fonts/DejaVuSans.ttf",
        "static/fonts/DejaVuSans.ttf",
    ]
    for p in candidates:
        try:
            pdf.add_font("DejaVu","",p,uni=True)
            pdf.add_font("DejaVu","B",p,uni=True)
            return True, p
        except Exception:
            continue
    return False, "not-found"

def generar_pdf_seguro(salida: Path, nombre_archivo: str, df_head: pd.DataFrame, resumen: pd.DataFrame, params: Dict[str,str]):
    try:
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        ok, _ = _try_add_dejavu(pdf)
        font = ("DejaVu","") if ok else ("helvetica","")
        font_b = ("DejaVu","B") if ok else ("helvetica","B")

        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        pdf.set_font(*font_b, size=18)
        pdf.cell(0, 10, "Mercury AI - Resumen de Reporte", ln=1)

        pdf.set_font(*font, size=11)
        pdf.cell(0, 8, f"Generado: {params.get('timestamp_iso')}", ln=1)
        pdf.cell(0, 8, f"Archivo: {nombre_archivo}", ln=1)

        pdf.set_font(*font_b, size=12)
        pdf.cell(0, 8, "Columnas (vista parcial):", ln=1)
        pdf.set_font(*font, size=10)
        pdf.multi_cell(0, 6, ", ".join(df_head.columns.tolist())[:160])

        pdf.set_font(*font_b, size=12)
        pdf.cell(0, 8, "Resumen estadístico (parcial):", ln=1)
        pdf.set_font(*font, size=9)
        head = resumen.head(12).fillna("").astype(str)
        headers = head.columns.tolist()
        pdf.cell(0, 6, " | ".join(headers)[:180], ln=1)
        for _, row in head.iterrows():
            pdf.cell(0, 5, " | ".join(row.values.tolist())[:180], ln=1)

        salida.parent.mkdir(parents=True, exist_ok=True)
        pdf.output(str(salida))
        return True, "ok"
    except Exception as e:
        return False, str(e)
