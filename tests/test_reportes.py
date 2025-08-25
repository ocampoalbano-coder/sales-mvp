# tests/test_reportes.py
from __future__ import annotations
import io
from pathlib import Path
from openpyxl import load_workbook
from app.core.procesamiento import (
    cargar_csv, limpiar_y_validar, ordenar_datos,
    agrupar_por_categoria, describir_dataframe
)
from app.core.generador_reportes import generar_excel

CSV1 = b"""ID,Nombre,Edad,Ingreso_Mensual,Fecha_Registro,Suscripcion_Activa,Categoria,Puntaje
1,Ana,34,4500,2022-04-05,True,A,72.3
2,Jose,41,7000,2021-11-10,False,B,66.1
3,Maria,29,0,02/03/2023,True,C,81.9
"""

def test_pipeline_excel_tmp(tmp_path: Path):
    df, meta = cargar_csv(io.BytesIO(CSV1))
    df2, val = limpiar_y_validar(df)
    df3 = ordenar_datos(df2)
    res = describir_dataframe(df3)
    cat = agrupar_por_categoria(df3)

    out = tmp_path / "rep.xlsx"
    generar_excel(
        df_original=df,
        df_limpio=df3,
        df_resumen=res,
        df_cat=cat,
        validaciones=val,
        parametros={"nombre_entrada":"mem","currency_symbol":"$"},
        currency_symbol="$",
        salida=out,
    )
    assert out.exists() and out.stat().st_size > 0
    wb = load_workbook(out)

    # Hojas clave
    assert {"datos_limpios","resumen","por_categoria","Validaciones","Parámetros"} <= set(wb.sheetnames)

    # 'por_categoria' contiene 'porcentaje' y fila TOTAL
    ws = wb["por_categoria"]
    headers = [c.value for c in ws[1]]
    assert "porcentaje" in headers
    vals = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]
    assert "TOTAL" in vals
